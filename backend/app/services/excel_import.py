"""
Excel 本体导入模板解析服务。

支持"对象信息 + 对象关系信息"两张表的 Excel 模板（如《XXX_本体导入模板.xlsx》），
解析为与 file_import.preview_json_ontology 完全一致的内存草稿结构
（objects / relations / actions / data_sources / summary），从而复用前端
"文件导入 → 专家审核 → 水合验证"三步流，无需新增落库逻辑。

设计要点（保证对后续新模板的兼容性）：
- 表头驱动：按"中文列名"定位列，而不是写死列位置。新模板即使调整列顺序、
  增删无关列，只要保留约定的列名即可正常导入。
- 容错：自动定位表头行、对象英文名前向填充、空行跳过。
"""
from io import BytesIO
from typing import Any

import openpyxl

# ── 属性类型映射（Excel 中文/英文类型 → 系统属性类型）──
_TYPE_MAP = {
    "string": "string", "varchar": "string", "char": "string", "text": "string",
    "字符串": "string", "文本": "string",
    "int": "number", "integer": "number", "number": "number", "decimal": "number",
    "double": "number", "float": "number", "long": "number", "bigint": "number",
    "数值": "number", "整数": "number", "数字": "number",
    "boolean": "boolean", "bool": "boolean", "布尔": "boolean",
    "date": "date", "datetime": "date", "timestamp": "date", "time": "date",
    "日期": "date", "时间": "date",
    "json": "json", "object": "json", "array": "json",
}

# ── 基数中文 → 系统基数 ──
_CARDINALITY_MAP = {
    "一对一": "1:1", "1:1": "1:1",
    "一对多": "1:N", "1:n": "1:N", "1:N": "1:N",
    "多对一": "N:1", "n:1": "N:1", "N:1": "N:1",
    "多对多": "N:N", "n:n": "N:N", "N:N": "N:N",
}

# ── 对象信息表：中文列名 → 字段键 ──
_OBJECT_COLS = {
    "对象英文名称": "obj_name",
    "对象中文名称": "obj_cn",
    "对象描述": "obj_desc",
    "状态": "obj_status",
    "属性英文名": "prop_name",
    "属性中文名": "prop_cn",
    "属性类型": "prop_type",
    "属性描述": "prop_desc",
    "是否主键": "prop_pk",
    "是否标题": "prop_title",
}

# ── 关系信息表：中文列名 → 字段键 ──
_RELATION_COLS = {
    "关系类型": "rel_type_cn",
    "对象": "source",
    "对象属性": "source_attr",
    "被关联对象": "target",
    "被关联对象属性": "target_attr",
    "关系标签1": "label1_cn",
    "关系标签1英文": "label1_en",
    "关系标签2": "label2_cn",
    "关系标签2英文": "label2_en",
    "基数设置": "cardinality",
}

_TRUE_VALUES = {"是", "y", "yes", "true", "1", "√", "✓"}


def _norm(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _is_true(v: Any) -> bool:
    return _norm(v).lower() in _TRUE_VALUES


def _map_type(raw: str) -> str:
    return _TYPE_MAP.get(raw.strip().lower(), "string") if raw else "string"


def _map_cardinality(raw: str) -> str:
    return _CARDINALITY_MAP.get(raw.strip(), "1:N") if raw else "1:N"


def _locate_header(ws, col_map: dict[str, str], scan_rows: int = 8) -> tuple[int, dict[str, int]]:
    """在前 scan_rows 行内定位表头行，返回 (表头行号, {字段键: 列索引})。

    表头行 = 命中 col_map 中已知列名最多的那一行。按列名文字定位列索引，
    因此容忍列顺序变化与未知新列。找不到返回 (-1, {})。
    """
    best_row, best_hits, best_idx = -1, 0, {}
    max_r = min(ws.max_row, scan_rows)
    for r in range(1, max_r + 1):
        idx: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            label = _norm(ws.cell(r, c).value)
            if label in col_map:
                idx[col_map[label]] = c
        if len(idx) > best_hits:
            best_row, best_hits, best_idx = r, len(idx), idx
    return best_row, best_idx


def _cell(ws, row: int, col_idx: dict[str, int], key: str) -> str:
    """按字段键取单元格文本，列不存在或为空时返回空串。"""
    c = col_idx.get(key)
    return _norm(ws.cell(row, c).value) if c else ""


def _parse_object_sheet(ws) -> tuple[list[dict], int]:
    """解析对象信息表，返回 (objects 草稿列表, 属性总数)。

    每个对象首行填对象元数据 + 首个属性；后续行只重复对象英文名（前向填充）+ 属性。
    """
    header_row, col_idx = _locate_header(ws, _OBJECT_COLS)
    if header_row == -1 or "obj_name" not in col_idx or "prop_name" not in col_idx:
        return [], 0

    objects: list[dict] = []
    by_name: dict[str, dict] = {}
    cur_name = ""  # 对象英文名前向填充
    prop_count = 0

    for r in range(header_row + 1, ws.max_row + 1):
        obj_name = _cell(ws, r, col_idx, "obj_name")
        if obj_name:
            cur_name = obj_name
        if not cur_name:
            continue

        # 新对象：首次出现时登记元数据
        if cur_name not in by_name:
            obj = {
                "name": cur_name,
                "display_name": _cell(ws, r, col_idx, "obj_cn") or cur_name,
                "tier": 2,
                "namespace": None,
                "primary_key": None,
                "description": _cell(ws, r, col_idx, "obj_desc"),
                "properties": [],
            }
            by_name[cur_name] = obj
            objects.append(obj)

        # 属性（同一行可能既有对象元数据又有首个属性）
        prop_name = _cell(ws, r, col_idx, "prop_name")
        if not prop_name:
            continue
        obj = by_name[cur_name]
        if any(p["name"] == prop_name for p in obj["properties"]):
            continue  # 同对象属性去重
        raw_type = _cell(ws, r, col_idx, "prop_type")
        obj["properties"].append({
            "name": prop_name,
            "display_name": _cell(ws, r, col_idx, "prop_cn") or prop_name,
            "type": _map_type(raw_type),
            "raw_type": raw_type or "string",
            "required": _is_true(_cell(ws, r, col_idx, "prop_pk")),
            "description": _cell(ws, r, col_idx, "prop_desc"),
            "source_table": None,
            "source_field": prop_name,
        })
        prop_count += 1
        # 主键 → 对象级 primary_key
        if _is_true(_cell(ws, r, col_idx, "prop_pk")) and not obj["primary_key"]:
            obj["primary_key"] = prop_name

    return objects, prop_count


def _slug(text: str, fallback: str) -> str:
    """把关系英文标签转成 snake_case 关系名，非 ASCII 或空则用 fallback。"""
    import re
    s = re.sub(r"[^a-zA-Z0-9]+", "_", text.strip().lower()).strip("_")
    return s or fallback


def _parse_relation_sheet(ws, valid_names: set[str]) -> list[dict]:
    """解析对象关系信息表，返回 relations 草稿列表（仅保留两端对象都存在的关系）。

    只映射核心字段：source/target/cardinality/关系名(标签1英文)/中文标签。
    中间数据集、双向标签、两端关联属性等按导入约定丢弃。
    """
    header_row, col_idx = _locate_header(ws, _RELATION_COLS)
    if header_row == -1 or "source" not in col_idx or "target" not in col_idx:
        return []

    relations: list[dict] = []
    for i, r in enumerate(range(header_row + 1, ws.max_row + 1)):
        source = _cell(ws, r, col_idx, "source")
        target = _cell(ws, r, col_idx, "target")
        if not source or not target:
            continue
        if source not in valid_names or target not in valid_names:
            continue  # 关系两端必须是已定义对象
        label_en = _cell(ws, r, col_idx, "label1_en")
        label_cn = _cell(ws, r, col_idx, "label1_cn")
        name = _slug(label_en, f"rel_{source}_{target}_{i}")
        relations.append({
            "name": name,
            "display_name": label_cn or label_en or name,
            "source": source,
            "target": target,
            "cardinality": _map_cardinality(_cell(ws, r, col_idx, "cardinality")),
            "description": label_cn or _cell(ws, r, col_idx, "rel_type_cn"),
        })
    return relations


def _classify_sheet(ws) -> str:
    """根据表头/标题判断 sheet 类型：'relation' 或 'object'。"""
    title = _norm(ws.title)
    if "关系" in title:
        return "relation"
    # 标题不含"关系"时，扫描前几行单元格判断
    for r in range(1, min(ws.max_row, 4) + 1):
        for c in range(1, ws.max_column + 1):
            if "被关联对象" in _norm(ws.cell(r, c).value):
                return "relation"
    return "object"


def preview_excel_ontology(content: bytes, namespace: str = "") -> dict:
    """解析 Excel 本体模板为草稿结构（结构与 preview_json_ontology 一致，不落库）。"""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)

    objects: list[dict] = []
    prop_count = 0
    relation_sheets: list[Any] = []

    for ws in wb.worksheets:
        if _classify_sheet(ws) == "relation":
            relation_sheets.append(ws)
        else:
            objs, pc = _parse_object_sheet(ws)
            objects.extend(objs)
            prop_count += pc

    valid_names = {o["name"] for o in objects}
    relations: list[dict] = []
    for ws in relation_sheets:
        relations.extend(_parse_relation_sheet(ws, valid_names))

    if namespace:
        for o in objects:
            o["namespace"] = namespace

    return {
        "objects": objects,
        "relations": relations,
        "actions": [],
        "data_sources": [],
        "summary": {
            "object_count": len(objects),
            "relation_count": len(relations),
            "property_count": prop_count,
            "action_count": 0,
        },
    }



