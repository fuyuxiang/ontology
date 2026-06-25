"""Excel 本体导入模板解析测试。"""
from io import BytesIO

import openpyxl

from app.services.excel_import import preview_excel_ontology


def _build_workbook() -> bytes:
    """构造一个与导入模板同构的最小 workbook。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对象信息"
    # 第1行注释占位，第2行分组占位，第3行表头
    ws["A1"] = "注：示例"
    ws["A2"] = "对象元数据信息"
    headers = ["对象英文名称", "对象中文名称", "对象描述", "状态",
               "属性英文名", "属性中文名", "属性类型", "属性描述", "是否主键", "是否标题"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    # User 对象：首行带元数据 + 首属性，后续行仅属性
    rows = [
        ["User", "用户", "用户对象", "启用", "user_id", "用户ID", "String", "主键", "是", "否"],
        [None, None, None, None, "name", "姓名", "String", "名称", None, "是"],
        [None, None, None, None, "age", "年龄", "Integer", "年龄", None, None],
        ["Order", "订单", "订单对象", "启用", "order_id", "订单ID", "String", "主键", "是", None],
        [None, None, None, None, "amount", "金额", "Decimal", "金额", None, None],
    ]
    for r, row in enumerate(rows, 4):
        for c, v in enumerate(row, 1):
            if v is not None:
                ws.cell(r, c, v)

    # 关系表
    ws2 = wb.create_sheet("对象关系信息")
    ws2["A1"] = "注：示例"
    rheaders = ["关系类型", "对象", "对象属性", "被关联对象", "被关联对象属性",
                "中间数据集", "中间数据集字段-对象", "中间数据集字段-被关联对象",
                "关系标签1", "关系标签1英文", "关系标签2", "关系标签2英文", "基数设置"]
    for c, h in enumerate(rheaders, 1):
        ws2.cell(2, c, h)  # 表头放第2行，测试表头自动定位
    ws2.append(["外键关联", "User", "user_id", "Order", "user_id",
                None, None, None, "下单", "places", "属于", "belongs to", "一对多"])
    # 指向不存在对象的关系应被跳过
    ws2.append(["外键关联", "User", "user_id", "Ghost", "x",
                None, None, None, "幽灵", "ghost", None, None, "一对多"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_preview_excel_basic():
    result = preview_excel_ontology(_build_workbook())

    assert result["summary"]["object_count"] == 2
    assert result["summary"]["property_count"] == 5
    assert result["summary"]["relation_count"] == 1  # Ghost 关系被跳过

    objs = {o["name"]: o for o in result["objects"]}
    assert objs["User"]["display_name"] == "用户"
    assert objs["User"]["primary_key"] == "user_id"
    assert len(objs["User"]["properties"]) == 3

    age = next(p for p in objs["User"]["properties"] if p["name"] == "age")
    assert age["type"] == "number"  # Integer → number
    assert age["display_name"] == "年龄"

    user_id = next(p for p in objs["User"]["properties"] if p["name"] == "user_id")
    assert user_id["required"] is True  # 主键必填

    rel = result["relations"][0]
    assert rel["source"] == "User"
    assert rel["target"] == "Order"
    assert rel["name"] == "places"  # 英文标签 slug
    assert rel["display_name"] == "下单"
    assert rel["cardinality"] == "1:N"  # 一对多


def test_preview_excel_header_reorder():
    """打乱对象表列顺序，仍应按列名正确解析。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对象信息"
    # 故意把属性英文名放在最前
    headers = ["属性英文名", "对象英文名称", "属性类型", "对象中文名称"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    ws.append(["uid", "Acct", "String", "账户"])
    buf = BytesIO()
    wb.save(buf)

    result = preview_excel_ontology(buf.getvalue())
    assert result["summary"]["object_count"] == 1
    obj = result["objects"][0]
    assert obj["name"] == "Acct"
    assert obj["display_name"] == "账户"
    assert obj["properties"][0]["name"] == "uid"


def test_preview_excel_empty():
    wb = openpyxl.Workbook()
    wb.active.title = "空表"
    buf = BytesIO()
    wb.save(buf)
    result = preview_excel_ontology(buf.getvalue())
    assert result["summary"]["object_count"] == 0
    assert result["relations"] == []
